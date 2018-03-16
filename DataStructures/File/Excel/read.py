# -*- coding: utf-8 -*-
'''读取 Excel 文档'''
import openpyxl


wb = openpyxl.load_workbook('example.xlsx')
print(type(wb)) # <class 'openpyxl.workbook.workbook.Workbook'>

#   从工作簿中取得工作表
wb.get_sheet_names() # ['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb.get_sheet_by_name('Sheet3')
print(sheet) # <Worksheet "Sheet3">
print(type(sheet)) # <class 'openpyxl.worksheet.worksheet.Worksheet'>
anotherSheet = wb.get_active_sheet()
print(anotherSheet) # <Worksheet "Sheet1">

#   从表中取得单元格
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet['A1']) # <Cell Sheet1.A1>
print(sheet['A1'].value) # datetime.datetime(2015, 4, 5, 13, 34, 2)
c = sheet['B1']
print(c.value) # 'Apples'
print(c.coordinate) # B1
print(sheet['C1'].value) # 73

sheet.cell(row=1, column=2) # <Cell Sheet1.B1>
print(sheet.cell(row=1, column=2).value) # 'Apples'
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

sheet.get_highest_row() # 7
sheet.get_highest_column() # 3

# 列字母和数字之间的转换
from openpyxl.cell import get_column_letter, column_index_from_string
get_column_letter(1) # 'A'
get_column_letter(2) # 'B'
get_column_letter(27) # 'AA'
get_column_letter(900) # 'AHP'
column_index_from_string('A') # 1
column_index_from_string('AA') # 27

#   从表中取得行和列
tuple(sheet['A1':'C3'])
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')
print(sheet.columns[1])
for cellObj in sheet.columns[1]:
    print(cellObj.value)