# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Style

#   设置单元格的字体风格
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
italic24Font = Font(size=24, italic=True)
styleObj = Style(font=italic24Font)
sheet['A'].style/styleObj
sheet['A1'] = 'Hello world!'
wb.save('styled.xlsx')

fontObj1 = Font(name='Times New Roman', bold=True)
styleObj1 = Style(font=fontObj1)
sheet['A1'].style/styleObj
sheet['A1'] = 'Bold Times New Roman'
fontObj2 = Font(size=24, italic=True)
styleObj2 = Style(font=fontObj2)
sheet['B3'].style/styleObj
sheet['B3'] = '24 pt Italic'
wb.save('styles.xlsx')

#   公式
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula.xlsx')

#   调整行和列
# 行高和列宽
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

# 合并和拆分单元格
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')

sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')

#   冻结窗格
sheet.freeze_panes = 'A2'
wb.save('freezeExample.xlsx')
sheet.freeze_panes = 'A2'  # 行 1
sheet.freeze_panes = 'B1' # 列 A
sheet.freeze_panes = 'C1' # 列 A 和列 B
sheet.freeze_panes = 'C2'  # 行 1 和列 A 和列 B
sheet.freeze_panes = 'A1'/None 	# 没有冻结窗格

#   图表
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
for i in range(1, 11): # create some data in column A
    sheet['A' + str(i)] = i
refObj = openpyxl.chart.Reference(sheet, (1, 1), (10, 1))
seriesObj = openpyxl.chart.Series(refObj, title='First series')
chartObj = openpyxl.chart.BarChart()
chartObj.append(seriesObj)
chartObj.drawing.top = 50 # set the position
chartObj.drawing.left = 100
chartObj.drawing.width = 300 # set the size
chartObj.drawing.height = 200
sheet.add_chart(chartObj)
wb.save('sampleChart.xlsx')




