# -*- coding: utf-8 -*-
import openpyxl

#   创建并保存 Excel 文档
wb = openpyxl.Workbook()
wb.get_sheet_names() # ['Sheet']
sheet = wb.get_active_sheet()
print(sheet.title) # 'Sheet'
sheet.title = 'Spam Bacon Eggs Sheet'
wb.get_sheet_names() # ['Spam Bacon Eggs Sheet']
wb.save('example_copy.xlsx')

#   创建和删除工作表: create_sheet() remove_sheet()
wb.create_sheet() # <Worksheet "Sheet1">
wb.get_sheet_names() # ['Sheet', 'Sheet1']
wb.create_sheet(index=0, title='First Sheet') # <Worksheet "First Sheet">
wb.get_sheet_names() # ['First Sheet', 'Sheet', 'Sheet1']
wb.create_sheet(index=2, title='Middle Sheet') # <Worksheet "Middle Sheet">
wb.get_sheet_names() # ['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']
# remove_sheet()方法接受一个Worksheet 对象作为其参数，而不是工作表名称的字符串
wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
wb.get_sheet_names() # ['First Sheet', 'Sheet']

#   将值写入单元格
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 'Hello world!'
print(sheet['A1'].value) # 'Hello world!